"""Excel export functionality for simulation results"""

from datetime import datetime
from io import BytesIO
from typing import List, Optional
import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import plotly.graph_objects as go
    import plotly.io as pio
    import kaleido
    KALEIDO_AVAILABLE = True
except ImportError:
    KALEIDO_AVAILABLE = False


def check_dependencies():
    """Check if required packages are installed"""
    missing = []
    if not OPENPYXL_AVAILABLE:
        missing.append("openpyxl")
    if not KALEIDO_AVAILABLE:
        missing.append("kaleido")
    return missing


def create_plotly_image(fig, width=800, height=500):
    """
    Export Plotly figure to PNG bytes for Excel embedding

    Args:
        fig: Plotly figure object
        width: Image width in pixels
        height: Image height in pixels

    Returns:
        BytesIO: PNG image data
    """
    if not KALEIDO_AVAILABLE:
        return None

    try:
        img_bytes = pio.to_image(fig, format='png', width=width, height=height)
        return BytesIO(img_bytes)
    except Exception as e:
        print(f"Error creating chart image: {e}")
        return None


def recreate_histogram(data, title, bins, hist_type, mean, median, xaxis_title="Value"):
    """
    Recreate histogram matching UI settings

    Args:
        data: List of values
        title: Chart title
        bins: Number of bins
        hist_type: 'Frequency' or 'Probability Density'
        mean: Mean value for vertical line
        median: Median value for vertical line
        xaxis_title: X-axis label

    Returns:
        BytesIO: PNG image or None
    """
    if not KALEIDO_AVAILABLE:
        return None

    hist_norm = 'probability density' if hist_type == 'Probability Density' else ''

    fig = go.Figure()
    fig.add_trace(go.Histogram(x=data, nbinsx=bins, histnorm=hist_norm, marker_color='lightblue'))
    fig.add_vline(x=mean, line_dash="dash", line_color="red", annotation_text="Mean")
    fig.add_vline(x=median, line_dash="dash", line_color="green", annotation_text="Median")
    fig.update_layout(
        title=title,
        xaxis_title=xaxis_title,
        yaxis_title=hist_type,
        showlegend=False,
        width=800,
        height=500,
        template="plotly_white"
    )

    return create_plotly_image(fig, width=800, height=500)


def format_header_row(ws, row, start_col=1, end_col=6, bg_color='366092', font_color='FFFFFF'):
    """Apply professional formatting to header rows"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name='Calibri', size=11, bold=True, color=font_color)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )


def add_section_title(ws, row, col, title, merge_cols=5):
    """Add formatted section title spanning multiple columns"""
    ws.cell(row=row, column=col, value=title)
    if merge_cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + merge_cols - 1)

    cell = ws.cell(row=row, column=col)
    cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')


def apply_alternating_rows(ws, start_row, end_row, start_col=1, end_col=6):
    """Apply alternating row colors for readability"""
    for row in range(start_row, end_row + 1):
        if (row - start_row) % 2 == 1:
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


def export_results_to_excel(
    alpha_summary,
    alpha_results,
    beta_diagnostics,
    beta_paths,
    gross_summary,
    gross_results,
    net_summary,
    net_results,
    beta_recon_diagnostics,
    decomp_diagnostics,
    config,
    hist_bins=50,
    hist_type='Frequency'
):
    """
    Main export function - creates complete Excel workbook

    Returns:
        BytesIO: Excel file ready for download
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()

    # Create sheets
    create_executive_summary(wb, alpha_summary, gross_summary, net_summary, config)
    create_alpha_sheet(wb, alpha_summary, alpha_results, hist_bins, hist_type)

    if beta_diagnostics and beta_paths is not None:
        create_beta_sheet(wb, beta_diagnostics, beta_paths)

    if gross_summary and gross_results:
        create_gross_sheet(wb, gross_summary, gross_results, beta_recon_diagnostics, beta_diagnostics, hist_bins, hist_type)

    if net_summary and net_results:
        create_net_sheet(wb, net_summary, net_results, hist_bins, hist_type)

    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_executive_summary(wb, alpha_summary, gross_summary, net_summary, config):
    """Create Executive Summary sheet"""
    ws = wb.active
    ws.title = "Executive Summary"

    row = 1

    # Header
    ws.cell(row=row, column=1, value="Monte Carlo Fund Simulation - Results Summary")
    ws.cell(row=row, column=1).font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    ws.cell(row=row, column=1, value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=row, column=1).font = Font(name='Calibri', size=10, italic=True)
    row += 2

    # Fund Information
    add_section_title(ws, row, 1, "Fund Information", merge_cols=6)
    row += 1

    ws.cell(row=row, column=1, value="Fund Name:")
    ws.cell(row=row, column=2, value=config.fund_name)
    row += 1

    ws.cell(row=row, column=1, value="Fund Manager:")
    ws.cell(row=row, column=2, value=config.fund_manager)
    row += 2

    # Simulation Parameters
    add_section_title(ws, row, 1, "Simulation Parameters", merge_cols=6)
    row += 1

    params = [
        ("Number of Simulations:", f"{config.simulation_count:,}"),
        ("Portfolio Size (Mean):", f"{config.investment_count_mean:.1f}"),
        ("Portfolio Size (Std Dev):", f"{config.investment_count_std:.1f}"),
        ("", ""),
        ("Leverage Rate:", f"{config.leverage_rate:.1%}"),
        ("Cost of Capital:", f"{config.cost_of_capital:.1%}"),
        ("Management Fee Rate:", f"{config.fee_rate:.1%}"),
        ("Carry Rate:", f"{config.carry_rate:.1%}"),
        ("Hurdle Rate:", f"{config.hurdle_rate:.1%}"),
        ("", ""),
        ("Beta Horizon (Trading Days):", f"{config.beta_horizon_days:,}"),
        ("Beta Paths:", f"{config.beta_n_paths:,}"),
        ("Beta Outlook:", config.beta_outlook.capitalize()),
        ("Beta Confidence:", config.beta_confidence.capitalize()),
    ]

    for label, value in params:
        if label:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Key Metrics Dashboard
    add_section_title(ws, row, 1, "Key Metrics Summary", merge_cols=6)
    row += 1

    # Headers
    headers = ["Metric", "Alpha", "Gross", "Net"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header)
    format_header_row(ws, row, 1, 4)
    row += 1

    # MOIC
    ws.cell(row=row, column=1, value="Median MOIC")
    ws.cell(row=row, column=2, value=f"{alpha_summary.median_moic:.2f}x")
    if gross_summary:
        ws.cell(row=row, column=3, value=f"{gross_summary.median_moic:.2f}x")
    if net_summary:
        ws.cell(row=row, column=4, value=f"{net_summary.median_moic:.2f}x")
    row += 1

    # IRR
    ws.cell(row=row, column=1, value="Median IRR")
    ws.cell(row=row, column=2, value=f"{alpha_summary.median_irr:.2%}")
    if gross_summary:
        ws.cell(row=row, column=3, value=f"{gross_summary.median_irr:.2%}")
    if net_summary:
        ws.cell(row=row, column=4, value=f"{net_summary.median_irr:.2%}")
    row += 1

    apply_alternating_rows(ws, row - 2, row - 1, 1, 4)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def create_alpha_sheet(wb, alpha_summary, alpha_results, hist_bins, hist_type):
    """Create Alpha Returns sheet"""
    ws = wb.create_sheet("Alpha Returns (Stage 1)")

    row = 1

    # Title
    add_section_title(ws, row, 1, "Stage 1: Alpha (Excess Returns)", merge_cols=6)
    row += 2

    # Statistics
    add_section_title(ws, row, 1, "MOIC Statistics", merge_cols=3)
    add_section_title(ws, row, 4, "IRR Statistics", merge_cols=3)
    row += 1

    # MOIC headers
    moic_headers = ["Metric", "Value"]
    for col, header in enumerate(moic_headers, start=1):
        ws.cell(row=row, column=col, value=header)

    # IRR headers
    irr_headers = ["Metric", "Value"]
    for col, header in enumerate(irr_headers, start=4):
        ws.cell(row=row, column=col, value=header)

    format_header_row(ws, row, 1, 2)
    format_header_row(ws, row, 4, 5)
    row += 1

    # MOIC stats
    moic_stats = [
        ("Mean", f"{alpha_summary.mean_moic:.2f}x"),
        ("Median", f"{alpha_summary.median_moic:.2f}x"),
        ("Std Dev", f"{alpha_summary.std_moic:.2f}x"),
        ("5th Percentile", f"{alpha_summary.percentile_5_moic:.2f}x"),
        ("95th Percentile", f"{alpha_summary.percentile_95_moic:.2f}x"),
    ]

    # IRR stats
    irr_stats = [
        ("Mean", f"{alpha_summary.mean_irr:.2%}"),
        ("Median", f"{alpha_summary.median_irr:.2%}"),
        ("Std Dev", f"{alpha_summary.std_irr:.2%}"),
        ("5th Percentile", f"{alpha_summary.percentile_5_irr:.2%}"),
        ("95th Percentile", f"{alpha_summary.percentile_95_irr:.2%}"),
    ]

    start_row = row
    for i, ((moic_label, moic_val), (irr_label, irr_val)) in enumerate(zip(moic_stats, irr_stats)):
        ws.cell(row=row, column=1, value=moic_label)
        ws.cell(row=row, column=2, value=moic_val)
        ws.cell(row=row, column=4, value=irr_label)
        ws.cell(row=row, column=5, value=irr_val)
        row += 1

    apply_alternating_rows(ws, start_row, row - 1, 1, 5)
    row += 2

    # Charts
    if KALEIDO_AVAILABLE:
        add_section_title(ws, row, 1, "Distribution Charts", merge_cols=6)
        row += 2

        # MOIC histogram
        moics = [r.moic for r in alpha_results]
        moic_img = recreate_histogram(
            moics,
            "Alpha MOIC Distribution",
            hist_bins,
            hist_type,
            alpha_summary.mean_moic,
            alpha_summary.median_moic,
            "Alpha MOIC"
        )

        if moic_img:
            img = XLImage(moic_img)
            img.width = 600
            img.height = 375
            ws.add_image(img, f'A{row}')
            row += 20

        # IRR histogram
        irrs = [r.irr * 100 for r in alpha_results]
        irr_img = recreate_histogram(
            irrs,
            "Alpha IRR Distribution",
            hist_bins,
            hist_type,
            alpha_summary.mean_irr * 100,
            alpha_summary.median_irr * 100,
            "Alpha IRR (%)"
        )

        if irr_img:
            img = XLImage(irr_img)
            img.width = 600
            img.height = 375
            ws.add_image(img, f'A{row}')

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15


def create_beta_sheet(wb, beta_diagnostics, beta_paths):
    """Create Beta Simulation sheet"""
    ws = wb.create_sheet("Beta Simulation (Stage 2)")

    row = 1

    # Title
    add_section_title(ws, row, 1, "Stage 2: Beta Forward Simulation", merge_cols=6)
    row += 2

    # Historical Data Metrics
    add_section_title(ws, row, 1, "Historical Data Metrics", merge_cols=6)
    row += 1

    hist_metrics = [
        ("Data Frequency:", beta_diagnostics.get('frequency', 'N/A').capitalize()),
        ("Observations:", f"{beta_diagnostics.get('n_observations', 0):,}"),
        ("Annual Return:", f"{beta_diagnostics.get('mu_hist_annual', 0):.2%}"),
        ("Annual Volatility:", f"{beta_diagnostics.get('sigma_hist_annual', 0):.2%}"),
    ]

    for label, value in hist_metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Forward Simulation Parameters
    add_section_title(ws, row, 1, "Forward Simulation Parameters", merge_cols=6)
    row += 1

    forward_params = [
        ("Outlook:", beta_diagnostics.get('outlook', 'N/A').capitalize()),
        ("Confidence:", beta_diagnostics.get('confidence', 'N/A').capitalize()),
        ("Target Return:", f"{beta_diagnostics.get('R_view', 0):.2%}"),
        ("Horizon (Trading Days):", f"{beta_diagnostics.get('horizon_days', 0):,}"),
        ("Paths Generated:", f"{beta_diagnostics.get('n_paths', 0):,}"),
    ]

    for label, value in forward_params:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Terminal Value Statistics
    if beta_paths is not None:
        add_section_title(ws, row, 1, "Terminal Value Statistics (Annualized Returns)", merge_cols=6)
        row += 1

        terminal_values = beta_paths.iloc[-1, :]
        start_price = beta_diagnostics.get('start_price', 1)
        trading_years = beta_diagnostics.get('horizon_days', 2520) / 252

        terminal_moics = terminal_values / start_price
        terminal_returns_annualized = (terminal_moics ** (1 / trading_years)) - 1

        terminal_stats = [
            ("Mean Return:", f"{terminal_returns_annualized.mean():.2%}"),
            ("Median Return:", f"{terminal_returns_annualized.median():.2%}"),
            ("5th Percentile:", f"{terminal_returns_annualized.quantile(0.05):.2%}"),
            ("95th Percentile:", f"{terminal_returns_annualized.quantile(0.95):.2%}"),
        ]

        for label, value in terminal_stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_gross_sheet(wb, gross_summary, gross_results, beta_recon_diagnostics, beta_diagnostics, hist_bins, hist_type):
    """Create Gross Performance sheet"""
    ws = wb.create_sheet("Gross Performance (Stage 3)")

    row = 1

    # Title
    add_section_title(ws, row, 1, "Stage 3: Gross Performance (Alpha × Beta)", merge_cols=6)
    row += 2

    # Beta Attribution (if available)
    if beta_recon_diagnostics and beta_diagnostics:
        if beta_recon_diagnostics.get('mean_beta_irr') is not None:
            add_section_title(ws, row, 1, "Beta Attribution Analysis", merge_cols=6)
            row += 1

            beta_target = beta_diagnostics.get('R_view', 0)
            actual_beta_irr = beta_recon_diagnostics.get('mean_beta_irr', 0)

            beta_attrs = [
                ("Target Beta Return:", f"{beta_target:.2%}"),
                ("Actual Mean Beta Return:", f"{actual_beta_irr:.2%}"),
                ("Beta IRR Range (5th-95th):",
                 f"{beta_recon_diagnostics.get('p5_beta_irr', 0):.2%} to {beta_recon_diagnostics.get('p95_beta_irr', 0):.2%}"),
                ("Investments Analyzed:", f"{beta_recon_diagnostics.get('n_investments', 0):,}"),
            ]

            for label, value in beta_attrs:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                row += 1

            row += 1

    # Statistics (same format as Alpha sheet)
    add_section_title(ws, row, 1, "MOIC Statistics", merge_cols=3)
    add_section_title(ws, row, 4, "IRR Statistics", merge_cols=3)
    row += 1

    # Headers
    for col, header in enumerate(["Metric", "Value"], start=1):
        ws.cell(row=row, column=col, value=header)
    for col, header in enumerate(["Metric", "Value"], start=4):
        ws.cell(row=row, column=col, value=header)

    format_header_row(ws, row, 1, 2)
    format_header_row(ws, row, 4, 5)
    row += 1

    # Stats
    moic_stats = [
        ("Mean", f"{gross_summary.mean_moic:.2f}x"),
        ("Median", f"{gross_summary.median_moic:.2f}x"),
        ("Std Dev", f"{gross_summary.std_moic:.2f}x"),
        ("5th Percentile", f"{gross_summary.percentile_5_moic:.2f}x"),
        ("95th Percentile", f"{gross_summary.percentile_95_moic:.2f}x"),
    ]

    irr_stats = [
        ("Mean", f"{gross_summary.mean_irr:.2%}"),
        ("Median", f"{gross_summary.median_irr:.2%}"),
        ("Std Dev", f"{gross_summary.std_irr:.2%}"),
        ("5th Percentile", f"{gross_summary.percentile_5_irr:.2%}"),
        ("95th Percentile", f"{gross_summary.percentile_95_irr:.2%}"),
    ]

    start_row = row
    for (moic_label, moic_val), (irr_label, irr_val) in zip(moic_stats, irr_stats):
        ws.cell(row=row, column=1, value=moic_label)
        ws.cell(row=row, column=2, value=moic_val)
        ws.cell(row=row, column=4, value=irr_label)
        ws.cell(row=row, column=5, value=irr_val)
        row += 1

    apply_alternating_rows(ws, start_row, row - 1, 1, 5)
    row += 2

    # Charts
    if KALEIDO_AVAILABLE:
        add_section_title(ws, row, 1, "Distribution Charts", merge_cols=6)
        row += 2

        # MOIC histogram
        moics = [r.moic for r in gross_results]
        moic_img = recreate_histogram(
            moics,
            "Gross MOIC Distribution",
            hist_bins,
            hist_type,
            gross_summary.mean_moic,
            gross_summary.median_moic,
            "MOIC"
        )

        if moic_img:
            img = XLImage(moic_img)
            img.width = 600
            img.height = 375
            ws.add_image(img, f'A{row}')
            row += 20

        # IRR histogram
        irrs = [r.irr * 100 for r in gross_results]
        irr_img = recreate_histogram(
            irrs,
            "Gross IRR Distribution",
            hist_bins,
            hist_type,
            gross_summary.mean_irr * 100,
            gross_summary.median_irr * 100,
            "IRR (%)"
        )

        if irr_img:
            img = XLImage(irr_img)
            img.width = 600
            img.height = 375
            ws.add_image(img, f'A{row}')

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15


def create_net_sheet(wb, net_summary, net_results, hist_bins, hist_type):
    """Create Net Performance sheet"""
    ws = wb.create_sheet("Net Performance (Stage 4)")

    row = 1

    # Title
    add_section_title(ws, row, 1, "Stage 4: Net Performance (After Costs)", merge_cols=6)
    row += 2

    # Cost Breakdown
    add_section_title(ws, row, 1, "Cost Breakdown (Averages)", merge_cols=6)
    row += 1

    avg_fees = sum(r.fees_paid for r in net_results) / len(net_results)
    avg_carry = sum(r.carry_paid for r in net_results) / len(net_results)
    avg_leverage = sum(r.leverage_cost for r in net_results) / len(net_results)
    avg_gross_profit = sum(r.gross_profit for r in net_results) / len(net_results)

    costs = [
        ("Avg Gross Profit:", f"${avg_gross_profit:,.0f}"),
        ("Avg Management Fees:", f"${avg_fees:,.0f}"),
        ("Avg Carry:", f"${avg_carry:,.0f}"),
        ("Avg Leverage Cost:", f"${avg_leverage:,.0f}"),
    ]

    for label, value in costs:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Statistics
    add_section_title(ws, row, 1, "MOIC Statistics", merge_cols=3)
    add_section_title(ws, row, 4, "IRR Statistics", merge_cols=3)
    row += 1

    # Headers
    for col, header in enumerate(["Metric", "Value"], start=1):
        ws.cell(row=row, column=col, value=header)
    for col, header in enumerate(["Metric", "Value"], start=4):
        ws.cell(row=row, column=col, value=header)

    format_header_row(ws, row, 1, 2)
    format_header_row(ws, row, 4, 5)
    row += 1

    # Stats
    moic_stats = [
        ("Mean", f"{net_summary.mean_moic:.2f}x"),
        ("Median", f"{net_summary.median_moic:.2f}x"),
        ("Std Dev", f"{net_summary.std_moic:.2f}x"),
        ("5th Percentile", f"{net_summary.percentile_5_moic:.2f}x"),
        ("95th Percentile", f"{net_summary.percentile_95_moic:.2f}x"),
    ]

    irr_stats = [
        ("Mean", f"{net_summary.mean_irr:.2%}"),
        ("Median", f"{net_summary.median_irr:.2%}"),
        ("Std Dev", f"{net_summary.std_irr:.2%}"),
        ("5th Percentile", f"{net_summary.percentile_5_irr:.2%}"),
        ("95th Percentile", f"{net_summary.percentile_95_irr:.2%}"),
    ]

    start_row = row
    for (moic_label, moic_val), (irr_label, irr_val) in zip(moic_stats, irr_stats):
        ws.cell(row=row, column=1, value=moic_label)
        ws.cell(row=row, column=2, value=moic_val)
        ws.cell(row=row, column=4, value=irr_label)
        ws.cell(row=row, column=5, value=irr_val)
        row += 1

    apply_alternating_rows(ws, start_row, row - 1, 1, 5)
    row += 2

    # Charts
    if KALEIDO_AVAILABLE:
        add_section_title(ws, row, 1, "Distribution Charts", merge_cols=6)
        row += 2

        # MOIC histogram
        moics = [r.moic for r in net_results]
        moic_img = recreate_histogram(
            moics,
            "Net MOIC Distribution",
            hist_bins,
            hist_type,
            net_summary.mean_moic,
            net_summary.median_moic,
            "MOIC"
        )

        if moic_img:
            img = XLImage(moic_img)
            img.width = 600
            img.height = 375
            ws.add_image(img, f'A{row}')
            row += 20

        # IRR histogram
        irrs = [r.irr * 100 for r in net_results]
        irr_img = recreate_histogram(
            irrs,
            "Net IRR Distribution",
            hist_bins,
            hist_type,
            net_summary.mean_irr * 100,
            net_summary.median_irr * 100,
            "IRR (%)"
        )

        if irr_img:
            img = XLImage(irr_img)
            img.width = 600
            img.height = 375
            ws.add_image(img, f'A{row}')

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
